from pandas import read_excel
from os import remove, listdir, mkdir, getcwd
from shutil import rmtree
from csv import reader
from openpyxl import load_workbook
from datetime import date
from time import sleep

def readExcelColumns():
    df = read_excel("controlGeneradores.xlsx", sheet_name="ControlGeneradores")
    indexColumnRBS = "RBS"
    indexColumnFechaAbastecimiento = "FECHA ULTIMO ABASTECIMIENTO"
    indexColumnFechaHorometro = "FECHA REGISTRO ULTIMO HOROMETRO"
    RBS = df.loc[:, indexColumnRBS]
    FECHAABASTECIMIENTO = df.loc[:, indexColumnFechaAbastecimiento]
    FECHAHOROMETRO = df.loc[:, indexColumnFechaHorometro]

def convertXLSX2CSV():
    # Convert controlGeneradores.xlsx to CSV
    df = read_excel("controlGeneradores.xlsx", sheet_name='ControlGeneradores')
    df.to_csv("auxControlGeneradores.csv", index = None, header = True, sep=";", encoding='latin1')
    # Convert all files in U2020 to CSV
    try: mkdir(getcwd()+"/auxFolder")
    except: pass
    for i in listdir("U2020/"):
        df = read_excel("U2020/"+i)
        df.to_csv("auxFolder/"+i[:-4]+"csv", index = None, header = True, sep=";", encoding='latin1')
        
def deleteAuxFiles():
    remove("auxControlGeneradores.csv")
    rmtree("auxFolder")

def time2Hours(time:str):
    # 3 hours 12 minutes 41 seconds
    try:
        hours = float(time[0:time.index(" hours")])
        minutes = float(time[time.index("hours")+6:time.index(" minutes")])
        seconds = float(time[time.index("minutes")+8:time.index(" seconds")])
        hours += minutes/60
        hours += seconds/3600
    except:
        return 0
    return hours

def next2Date(nextDate:str, day:int, month:int, year:int):
    if year < int(nextDate[-4:]):
        return True
    if year == int(nextDate[-4:]):
        if month < int(nextDate[-7:-5]):
            return True
        elif month == int(nextDate[-7:-5]):
            if day <= int(nextDate[:2]):
                return True
    return False

def before2Date(beforeDate:str, day:int, month:int, year:int):
    if year > int(beforeDate[-4:]):
        return True
    if year == int(beforeDate[-4:]):
        if month > int(beforeDate[-7:-5]):
            return True
        elif month == int(beforeDate[-7:-5]):
            if day > int(beforeDate[:2]):
                return True
    return False

def hoursGenWorkSinceDate(RBS:str, day:int, month:int, year:int) -> float:
    hours = 0
    dateFirst = []
    for i in listdir("auxFolder/"):
        with open("auxFolder/"+i) as csvU2020:
            U2020 = reader(csvU2020)
            for row in U2020:
                # RBS Found
                RBSFlat = str(row[0]).split(";")[6]
                if RBSFlat[:2].upper() == "R1" or RBSFlat[:2].upper() == "R2":
                    RBSFlat = RBSFlat[5:]
                if RBSFlat[-3:].upper() == "W08":
                    RBSFlat = RBSFlat[:-3]
                if RBSFlat.upper() == RBS.upper() and next2Date(str(row[0]).split(";")[8][:10], day, month, year) and (str(row[0]).split(";")[7] == "PWR GENERADOR_ENCENDIDO" or str(row[0]).split(";")[7] == "PWR GENERADOR_ENCENDIDO_REPETIDOR_MW"):
                    hours += time2Hours(str(row[0]).split(";")[10])
                    if dateFirst == []:
                        dateFirst = [int(str(row[0]).split(";")[8][:10][:2]), int(str(row[0]).split(";")[8][:10][-7:-5]), int(str(row[0]).split(";")[8][:10][-4:])]
                    else:
                        dateFirst = previousDate([int(str(row[0]).split(";")[8][:10][:2]), int(str(row[0]).split(";")[8][:10][-7:-5]), int(str(row[0]).split(";")[8][:10][-4:])], dateFirst)
    return hours, dateFirst

def previousDate(date1:list[int], date2:list[int]):
    # day, month, year
    if date1[2]<date2[2]:
        return date1
    elif date1[2]>date2[2]:
        return date2
    else:
        if date1[1]<date2[1]:
            return date1
        elif date1[1]>date2[1]:
            return date2
        else:
            if date1[0]<date2[0]:
                return date1
            elif date1[0]>date2[0]:
                return date2
            else: return date1

def calculateHoursGenerator():
    workbook = load_workbook("controlGeneradores.xlsx")
    worksheet = workbook.active
    bandera = 0
    with open("auxControlGeneradores.csv") as csvControlGeneradores:
        controlGeneradores = reader(csvControlGeneradores)
        for row in controlGeneradores:
            bandera +=1
            hoursUltimoAbastecimiento = 0
            hoursUltimoHorometro = 0
            hoursUltimoCambioAceite = 0
            date1 = ["", "", ""]
            date2 = ["", "", ""]
            date3 = ["", "", ""]
            RBS = str(row[0]).split(";")[0]
            #print(RBS)
            if RBS != "RBS" and RBS != "(gal)\"":
                try: hoursUltimoAbastecimiento, date1 = hoursGenWorkSinceDate(RBS = RBS, day = int(str(row[0]).split(";")[13].split(" ")[0].split("-")[2]), month = int(str(row[0]).split(";")[13].split(" ")[0].split("-")[1]), year = int(str(row[0]).split(";")[13].split(" ")[0].split("-")[0]))
                except Exception as e: pass
                try:hoursUltimoHorometro, date2 = hoursGenWorkSinceDate(RBS = RBS, day = int(str(row[0]).split(";")[17].split(" ")[0].split("-")[2]), month = int(str(row[0]).split(";")[17].split(" ")[0].split("-")[1]), year = int(str(row[0]).split(";")[17].split(" ")[0].split("-")[0]))
                except Exception as e: pass
                try:hoursUltimoCambioAceite, date3 = hoursGenWorkSinceDate(RBS = RBS, day = int(str(row[0]).split(";")[20].split(" ")[0].split("-")[2]), month = int(str(row[0]).split(";")[20].split(" ")[0].split("-")[1]), year = int(str(row[0]).split(";")[20].split(" ")[0].split("-")[0]))
                except Exception as e: pass
                print(RBS, "- Horas Ultimo Abastecimiento:", hoursUltimoAbastecimiento.__round__(2))
                print(RBS, "- Horas Ultimo Horometro:", hoursUltimoHorometro.__round__(2))
                print(RBS, "- Horas Ultimo Cambio Aceite:", hoursUltimoCambioAceite.__round__(2))
                
                # Writing the xlsx
                try:worksheet['AA' + str(bandera-1)].value = str(date1[0]) + "-" + str(date1[1]) + "-" + str(date1[2])
                except: worksheet['AA' + str(bandera-1)].value = ""
                worksheet['AB' + str(bandera-1)].value = str(hoursUltimoAbastecimiento.__round__(2))
                
                try:worksheet['AC' + str(bandera-1)].value = str(date3[0]) + "-" + str(date3[1]) + "-" + str(date3[2])
                except: worksheet['AC' + str(bandera-1)].value = ""
                worksheet['AD' + str(bandera-1)].value = str(hoursUltimoCambioAceite.__round__(2))

                try:worksheet['AE' + str(bandera-1)].value = str(date2[0]) + "-" + str(date2[1]) + "-" + str(date2[2])
                except: worksheet['AE' + str(bandera-1)].value = ""
                worksheet['AF' + str(bandera-1)].value = str(hoursUltimoHorometro.__round__(2))
    workbook.save("controlGeneradores.xlsx")
    workbook.close   

def hoursGenWorkSinceDate2Date(RBS, fromDate, untilDate):
    hours = 0
    for i in listdir("auxFolder/"):
        with open("auxFolder/"+i) as csvU2020:
            U2020 = reader(csvU2020)
            for row in U2020:
                # RBS Found
                RBSFlat = str(row[0]).split(";")[6]
                if RBSFlat[:2].upper() == "R1" or RBSFlat[:2].upper() == "R2":
                    RBSFlat = RBSFlat[5:]
                if RBSFlat[-3:].upper() == "W08":
                    RBSFlat = RBSFlat[:-3]
                if RBSFlat.upper() == RBS.upper() and next2Date(str(row[0]).split(";")[8][:10], fromDate[0], fromDate[1], fromDate[2]) and str(row[0]).split(";")[7] == "PWR GENERADOR_ENCENDIDO":
                    if before2Date(str(row[0]).split(";")[8][:10], untilDate[0], untilDate[1], untilDate[2]):
                        hours += time2Hours(str(row[0]).split(";")[10])
    return hours
    pass

def calculateMeanHours(fromDate, untilDate):
    workbook = load_workbook("controlGeneradores.xlsx")
    worksheet = workbook.active
    bandera = 0
    with open("auxControlGeneradores.csv") as csvControlGeneradores:
        controlGeneradores = reader(csvControlGeneradores)
        for row in controlGeneradores:
            bandera +=1
            RBS = str(row[0]).split(";")[0]
            if RBS != "RBS" and RBS != "(gal)\"":
                try: hours = hoursGenWorkSinceDate2Date(RBS = RBS, fromDate = fromDate, untilDate = untilDate)
                except: pass
                d0 = date(fromDate[2], fromDate[1], fromDate[0])
                d1 = date(untilDate[2], untilDate[1], untilDate[0])
                delta = d1 - d0
                print(RBS, "- Horas/Dia Trabajadas por el generador:", (hours/int(delta.days)).__round__(2))
                # Writing the xlsx
                worksheet['AJ' + str(bandera-1)].value = (hours/int(delta.days)).__round__(2)
    print("Calculado las horas de trabajo en los "+str(delta.days)+" dias solicitados")
    workbook.save("controlGeneradores.xlsx")
    workbook.close  

def main():
    opcion = 0
    while opcion != 3:
        print("1. Calcular horas encendidas del generador (Desde ultimo abastecimiento, Cambio de aceite y Horometro)")
        print("2. Calcular promedio de horas trabajadas por el generador")
        print("3. Salir")
        while opcion < 1 or opcion > 3:
            try:
                opcion = int(input("Seleccione la accion a realizar: "))
            except:
                print("Opcion incorrecta.")
        if opcion == 1:
            convertXLSX2CSV()
            calculateHoursGenerator()
            deleteAuxFiles()
            print("Horas calculadas")
            opcion = 0
        elif opcion == 2:
            convertXLSX2CSV()
            print("Ingrese desde que fecha se va a hacer el calculo")
            fromDay = int(input("Dia: "))
            fromMonth = int(input("Mes: "))
            fromYear = int(input("Año: "))
            print("Ingrese hasta que fecha se va a hacer el calculo")
            untilDay = int(input("Dia: "))+1
            untilMonth = int(input("Mes: "))
            untilYear = int(input("Año: "))
            calculateMeanHours([fromDay, fromMonth, fromYear], [untilDay, untilMonth, untilYear])
            deleteAuxFiles()
            print("Horas calculadas")
            opcion = 0
        else:
            print("Saliendo del programa")

if __name__ == "__main__":
    main()